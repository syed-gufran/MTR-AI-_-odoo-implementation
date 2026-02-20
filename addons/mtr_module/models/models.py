# -*- coding: utf-8 -*-
import base64
import csv
import datetime
import io
import json
import re

from odoo import _, api, fields, models, tools
from odoo.exceptions import UserError, ValidationError


def _normalize_header(value):
    normalized = re.sub(r"[^a-z0-9]+", "_", (value or "").strip().lower())
    return normalized.strip("_")


def _to_float(value):
    if value in (None, ""):
        return False
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).strip().replace(",", "")
    if not cleaned:
        return False
    try:
        return float(cleaned)
    except ValueError:
        return False


def _to_date(value):
    if value in (None, "", False):
        return False
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d-%m-%Y"):
        try:
            return datetime.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return False


def _to_int(value):
    number = _to_float(value)
    if number is False:
        return False
    return int(number)


class MtrData(models.Model):
    _name = "mtr.data"
    _description = "MTR Data"
    _rec_name = "heat_number"
    _order = "id desc"
    _sql_constraints = [
        (
            "mtr_heat_batch_unique",
            "unique(heat_number, batch_number)",
            "An MTR with the same heat number and batch number already exists.",
        )
    ]

    heat_number = fields.Char(required=True, index=True)
    batch_number = fields.Char(required=True)
    grade = fields.Char()
    manufacturer = fields.Char()
    certificate_number = fields.Char()
    certificate_date = fields.Date()

    c_element = fields.Float(string="C")
    mn_element = fields.Float(string="Mn")
    si_element = fields.Float(string="Si")
    p_element = fields.Float(string="P")
    s_element = fields.Float(string="S")
    cu_element = fields.Float(string="Cu")
    ni_element = fields.Float(string="Ni")
    cr_element = fields.Float(string="Cr")
    mo_element = fields.Float(string="Mo")
    n_element = fields.Float(string="N")

    yield_strength = fields.Float()
    tensile_strength = fields.Float()
    elongation = fields.Float()
    reduction_area = fields.Float()
    hardness = fields.Float()

    impact_test_temp = fields.Float()
    impact_coupon_size = fields.Char()
    impact_specimen_1 = fields.Float()
    impact_specimen_2 = fields.Float()
    impact_specimen_3 = fields.Float()
    impact_average = fields.Float()

    country_of_melt = fields.Char()
    country_of_manufacture = fields.Char()

    source_file = fields.Char()
    uploaded_at = fields.Datetime(default=fields.Datetime.now)

    @api.constrains("heat_number", "batch_number")
    def _check_required_keys(self):
        for record in self:
            if not record.heat_number or not record.batch_number:
                raise ValidationError(_("Heat Number and Batch Number are required."))

    @api.model
    def upsert_from_payload(self, payload):
        if not isinstance(payload, dict):
            raise UserError(_("Payload must be a dictionary."))

        heat_number = (payload.get("heat_number") or "").strip()
        batch_number = (payload.get("batch_number") or "").strip()
        if not heat_number or not batch_number:
            raise UserError(_("heat_number and batch_number are required."))

        values = {
            "heat_number": heat_number,
            "batch_number": batch_number,
            "grade": payload.get("grade"),
            "manufacturer": payload.get("manufacturer"),
            "certificate_number": payload.get("certificate_number"),
            "certificate_date": _to_date(payload.get("certificate_date")),
            "c_element": _to_float(payload.get("c_element") or payload.get("c")),
            "mn_element": _to_float(payload.get("mn_element") or payload.get("mn")),
            "si_element": _to_float(payload.get("si_element") or payload.get("si")),
            "p_element": _to_float(payload.get("p_element") or payload.get("p")),
            "s_element": _to_float(payload.get("s_element") or payload.get("s")),
            "cu_element": _to_float(payload.get("cu_element") or payload.get("cu")),
            "ni_element": _to_float(payload.get("ni_element") or payload.get("ni")),
            "cr_element": _to_float(payload.get("cr_element") or payload.get("cr")),
            "mo_element": _to_float(payload.get("mo_element") or payload.get("mo")),
            "n_element": _to_float(payload.get("n_element") or payload.get("n")),
            "yield_strength": _to_float(payload.get("yield_strength")),
            "tensile_strength": _to_float(payload.get("tensile_strength")),
            "elongation": _to_float(payload.get("elongation")),
            "reduction_area": _to_float(payload.get("reduction_area")),
            "hardness": _to_float(payload.get("hardness")),
            "impact_test_temp": _to_float(payload.get("impact_test_temp")),
            "impact_coupon_size": payload.get("impact_coupon_size"),
            "impact_specimen_1": _to_float(payload.get("impact_specimen_1")),
            "impact_specimen_2": _to_float(payload.get("impact_specimen_2")),
            "impact_specimen_3": _to_float(payload.get("impact_specimen_3")),
            "impact_average": _to_float(payload.get("impact_average")),
            "country_of_melt": payload.get("country_of_melt"),
            "country_of_manufacture": payload.get("country_of_manufacture"),
            "source_file": payload.get("source_file"),
            "uploaded_at": fields.Datetime.now(),
        }

        existing = self.search(
            [("heat_number", "=", heat_number), ("batch_number", "=", batch_number)],
            limit=1,
        )
        if existing:
            existing.write(values)
            return {"id": existing.id, "operation": "updated"}
        created = self.create(values)
        return {"id": created.id, "operation": "created"}


class InventoryRecord(models.Model):
    _name = "inventory.record"
    _description = "Inventory Record"
    _rec_name = "lot_number"
    _order = "id desc"
    _table = "inventory"

    date = fields.Date()
    location_code = fields.Char()
    item_no = fields.Char(index=True)
    quantity = fields.Float()
    unit_of_measure_code = fields.Char()
    document_no = fields.Char()
    wsi_variant_code = fields.Char()
    dimensions = fields.Char()
    lot_number = fields.Char(index=True)
    heat_number = fields.Char(index=True)
    slab_number = fields.Char()
    internal_bin = fields.Char()
    additional_notes = fields.Text()
    cost_amount_actual = fields.Float()
    description_2 = fields.Char()
    origin_code = fields.Char()
    picked = fields.Char()
    cutting_plan_no = fields.Char()
    image_path = fields.Char()
    entry_type = fields.Char()
    document_type = fields.Char()
    drawing = fields.Char()
    yield_value = fields.Float(string="Yield")
    document_line_no = fields.Integer()
    revision = fields.Char()
    laser_quality = fields.Char()
    unitcost_cwt = fields.Float(string="UnitCost / CWT")
    piece_no = fields.Char()
    variant_code = fields.Char()
    description = fields.Char()
    return_reason_code = fields.Char()
    serial_no = fields.Char()
    package_no = fields.Char()
    invoiced_quantity = fields.Float()
    inventory_by_location = fields.Float()
    inventory = fields.Float()
    expiration_date = fields.Date()
    remaining_quantity = fields.Float()
    shipped_qty_not_returned = fields.Float()
    reserved_quantity = fields.Float()
    qty_per_unit_of_measure = fields.Float()
    sales_amount_expected = fields.Float()
    sales_amount_actual = fields.Float()
    cost_amount_expected = fields.Float()
    cost_amount_non_invtbl = fields.Float()
    item_description = fields.Char()
    cost_amount_expected_acy = fields.Float()
    cost_amount_actual_acy = fields.Float()
    completely_invoiced = fields.Char()
    cost_amount_non_invtbl_acy = fields.Float()
    assemble_to_order = fields.Char()
    drop_shipment = fields.Char()
    open_flag = fields.Char(string="Open")
    order_type = fields.Char()
    order_no = fields.Char()
    order_line_no = fields.Integer()
    prod_order_comp_line_no = fields.Integer()
    entry_no = fields.Integer(index=True)
    project_no = fields.Char()
    project_task_no = fields.Char()
    source_type = fields.Char()
    source_no = fields.Char()
    source_description = fields.Char()
    source_order_no = fields.Char()

    grade = fields.Char()
    weight = fields.Float()
    posting_date = fields.Date()
    country_of_melt = fields.Char()
    country_of_manufacture = fields.Char()
    source_file = fields.Char()
    raw_row_data = fields.Text(help="Full BC row as JSON for unmapped columns.")


class InventoryImportWizard(models.TransientModel):
    _name = "inventory.import.wizard"
    _description = "Inventory Import Wizard"

    file_data = fields.Binary(required=True)
    file_name = fields.Char(required=True)
    delimiter = fields.Char(default=",")
    has_header = fields.Boolean(default=True)

    def action_import(self):
        self.ensure_one()
        if not self.file_data or not self.file_name:
            raise UserError(_("Please upload a file."))

        rows = self._read_rows()
        if not rows:
            raise UserError(_("No data rows were found in the file."))

        Inventory = self.env["inventory.record"].sudo()
        Inventory.search([]).unlink()

        to_create = []
        for row in rows:
            normalized = {_normalize_header(k): v for k, v in row.items()}
            to_create.append(self._map_inventory_row(normalized, row))

        Inventory.create(to_create)
        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("Import Complete"),
                "message": _("Imported %s inventory rows.") % len(to_create),
                "type": "success",
                "sticky": False,
            },
        }

    def _map_inventory_row(self, normalized, row):
        date_value = _to_date(normalized.get("date"))
        return {
            "date": date_value,
            "location_code": normalized.get("location_code"),
            "item_no": normalized.get("item_no"),
            "quantity": _to_float(normalized.get("quantity")),
            "unit_of_measure_code": normalized.get("unit_of_measure_code"),
            "document_no": normalized.get("document_no"),
            "wsi_variant_code": normalized.get("wsi_variant_code"),
            "dimensions": normalized.get("dimensions"),
            "lot_number": normalized.get("lot_no") or normalized.get("lot_number"),
            "heat_number": normalized.get("heat_no") or normalized.get("heat_number"),
            "slab_number": normalized.get("slab_no") or normalized.get("slab_number"),
            "internal_bin": normalized.get("internal_bin"),
            "additional_notes": normalized.get("additional_notes"),
            "cost_amount_actual": _to_float(normalized.get("cost_amount_actual")),
            "description_2": normalized.get("description_2"),
            "origin_code": normalized.get("origin_code"),
            "picked": str(normalized.get("picked")) if normalized.get("picked") not in (None, "") else False,
            "cutting_plan_no": normalized.get("cutting_plan_no"),
            "image_path": normalized.get("image_path"),
            "entry_type": normalized.get("entry_type"),
            "document_type": normalized.get("document_type"),
            "drawing": normalized.get("drawing"),
            "yield_value": _to_float(normalized.get("yield")),
            "document_line_no": _to_int(normalized.get("document_line_no")),
            "revision": normalized.get("revision"),
            "laser_quality": normalized.get("laser_quality"),
            "unitcost_cwt": _to_float(normalized.get("unitcost_cwt")),
            "piece_no": normalized.get("piece_no"),
            "variant_code": normalized.get("variant_code"),
            "description": normalized.get("description"),
            "return_reason_code": normalized.get("return_reason_code"),
            "serial_no": normalized.get("serial_no"),
            "package_no": normalized.get("package_no"),
            "invoiced_quantity": _to_float(normalized.get("invoiced_quantity")),
            "inventory_by_location": _to_float(normalized.get("inventory_by_location")),
            "inventory": _to_float(normalized.get("inventory")),
            "expiration_date": _to_date(normalized.get("expiration_date")),
            "remaining_quantity": _to_float(normalized.get("remaining_quantity")),
            "shipped_qty_not_returned": _to_float(normalized.get("shipped_qty_not_returned")),
            "reserved_quantity": _to_float(normalized.get("reserved_quantity")),
            "qty_per_unit_of_measure": _to_float(normalized.get("qty_per_unit_of_measure")),
            "sales_amount_expected": _to_float(normalized.get("sales_amount_expected")),
            "sales_amount_actual": _to_float(normalized.get("sales_amount_actual")),
            "cost_amount_expected": _to_float(normalized.get("cost_amount_expected")),
            "cost_amount_non_invtbl": _to_float(normalized.get("cost_amount_non_invtbl")),
            "item_description": normalized.get("item_description"),
            "cost_amount_expected_acy": _to_float(normalized.get("cost_amount_expected_acy")),
            "cost_amount_actual_acy": _to_float(normalized.get("cost_amount_actual_acy")),
            "completely_invoiced": str(normalized.get("completely_invoiced"))
            if normalized.get("completely_invoiced") not in (None, "")
            else False,
            "cost_amount_non_invtbl_acy": _to_float(normalized.get("cost_amount_non_invtbl_acy")),
            "assemble_to_order": str(normalized.get("assemble_to_order"))
            if normalized.get("assemble_to_order") not in (None, "")
            else False,
            "drop_shipment": str(normalized.get("drop_shipment"))
            if normalized.get("drop_shipment") not in (None, "")
            else False,
            "open_flag": str(normalized.get("open")) if normalized.get("open") not in (None, "") else False,
            "order_type": normalized.get("order_type"),
            "order_no": normalized.get("order_no"),
            "order_line_no": _to_int(normalized.get("order_line_no")),
            "prod_order_comp_line_no": _to_int(normalized.get("prod_order_comp_line_no")),
            "entry_no": _to_int(normalized.get("entry_no")),
            "project_no": normalized.get("project_no"),
            "project_task_no": normalized.get("project_task_no"),
            "source_type": normalized.get("source_type"),
            "source_no": normalized.get("source_no"),
            "source_description": normalized.get("source_description"),
            "source_order_no": normalized.get("source_order_no"),
            "grade": normalized.get("grade"),
            "weight": _to_float(normalized.get("weight")),
            "posting_date": _to_date(normalized.get("posting_date")) or date_value,
            "country_of_melt": normalized.get("country_of_melt"),
            "country_of_manufacture": normalized.get("country_of_manufacture"),
            "source_file": self.file_name,
            "raw_row_data": json.dumps(row, default=str),
        }

    def _read_rows(self):
        file_name = (self.file_name or "").lower()
        data = base64.b64decode(self.file_data)
        if file_name.endswith(".csv"):
            return self._read_csv(data)
        if file_name.endswith(".xlsx"):
            return self._read_xlsx(data)
        raise UserError(_("Only CSV and XLSX files are supported."))

    def _read_csv(self, data):
        text = data.decode("utf-8-sig")
        buffer = io.StringIO(text)
        reader = csv.DictReader(buffer, delimiter=self.delimiter or ",")
        return [row for row in reader if any(row.values())]

    def _read_xlsx(self, data):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise UserError(_("openpyxl is required for XLSX imports."))

        workbook = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        payload = []
        for line in rows[1:]:
            if not any(line):
                continue
            row = {headers[idx]: (line[idx] if idx < len(line) else "") for idx in range(len(headers))}
            payload.append(row)
        return payload


class MtrInventoryJoinReport(models.Model):
    _name = "mtr.inventory.join.report"
    _description = "MTR Inventory Join Report"
    _auto = False
    _rec_name = "inv_lot_number"

    join_status = fields.Char(readonly=True)

    inv_entry_no = fields.Integer(readonly=True)
    inv_date = fields.Date(readonly=True)
    inv_location_code = fields.Char(readonly=True)
    inv_item_no = fields.Char(readonly=True)
    inv_quantity = fields.Float(readonly=True)
    inv_unit_of_measure_code = fields.Char(readonly=True)
    inv_document_no = fields.Char(readonly=True)
    inv_wsi_variant_code = fields.Char(readonly=True)
    inv_dimensions = fields.Char(readonly=True)
    inv_lot_number = fields.Char(readonly=True)
    inv_heat_number = fields.Char(readonly=True)
    inv_slab_number = fields.Char(readonly=True)
    inv_internal_bin = fields.Char(readonly=True)
    inv_additional_notes = fields.Text(readonly=True)
    inv_cost_amount_actual = fields.Float(readonly=True)
    inv_description_2 = fields.Char(readonly=True)
    inv_origin_code = fields.Char(readonly=True)
    inv_picked = fields.Char(readonly=True)
    inv_cutting_plan_no = fields.Char(readonly=True)
    inv_image_path = fields.Char(readonly=True)
    inv_entry_type = fields.Char(readonly=True)
    inv_document_type = fields.Char(readonly=True)
    inv_drawing = fields.Char(readonly=True)
    inv_yield_value = fields.Float(readonly=True)
    inv_document_line_no = fields.Integer(readonly=True)
    inv_revision = fields.Char(readonly=True)
    inv_laser_quality = fields.Char(readonly=True)
    inv_unitcost_cwt = fields.Float(readonly=True)
    inv_piece_no = fields.Char(readonly=True)
    inv_variant_code = fields.Char(readonly=True)
    inv_description = fields.Char(readonly=True)
    inv_return_reason_code = fields.Char(readonly=True)
    inv_serial_no = fields.Char(readonly=True)
    inv_package_no = fields.Char(readonly=True)
    inv_invoiced_quantity = fields.Float(readonly=True)
    inv_inventory_by_location = fields.Float(readonly=True)
    inv_inventory = fields.Float(readonly=True)
    inv_expiration_date = fields.Date(readonly=True)
    inv_remaining_quantity = fields.Float(readonly=True)
    inv_shipped_qty_not_returned = fields.Float(readonly=True)
    inv_reserved_quantity = fields.Float(readonly=True)
    inv_qty_per_unit_of_measure = fields.Float(readonly=True)
    inv_sales_amount_expected = fields.Float(readonly=True)
    inv_sales_amount_actual = fields.Float(readonly=True)
    inv_cost_amount_expected = fields.Float(readonly=True)
    inv_cost_amount_non_invtbl = fields.Float(readonly=True)
    inv_item_description = fields.Char(readonly=True)
    inv_cost_amount_expected_acy = fields.Float(readonly=True)
    inv_cost_amount_actual_acy = fields.Float(readonly=True)
    inv_completely_invoiced = fields.Char(readonly=True)
    inv_cost_amount_non_invtbl_acy = fields.Float(readonly=True)
    inv_assemble_to_order = fields.Char(readonly=True)
    inv_drop_shipment = fields.Char(readonly=True)
    inv_open_flag = fields.Char(readonly=True)
    inv_order_type = fields.Char(readonly=True)
    inv_order_no = fields.Char(readonly=True)
    inv_order_line_no = fields.Integer(readonly=True)
    inv_prod_order_comp_line_no = fields.Integer(readonly=True)
    inv_project_no = fields.Char(readonly=True)
    inv_project_task_no = fields.Char(readonly=True)
    inv_source_type = fields.Char(readonly=True)
    inv_source_no = fields.Char(readonly=True)
    inv_source_description = fields.Char(readonly=True)
    inv_source_order_no = fields.Char(readonly=True)
    inv_grade = fields.Char(readonly=True)
    inv_weight = fields.Float(readonly=True)
    inv_posting_date = fields.Date(readonly=True)
    inv_country_of_melt = fields.Char(readonly=True)
    inv_country_of_manufacture = fields.Char(readonly=True)
    inv_source_file = fields.Char(readonly=True)

    mtr_id = fields.Integer(readonly=True)
    mtr_heat_number = fields.Char(readonly=True)
    mtr_batch_number = fields.Char(readonly=True)
    mtr_grade = fields.Char(readonly=True)
    mtr_manufacturer = fields.Char(readonly=True)
    mtr_certificate_number = fields.Char(readonly=True)
    mtr_certificate_date = fields.Date(readonly=True)
    mtr_c = fields.Float(readonly=True)
    mtr_mn = fields.Float(readonly=True)
    mtr_si = fields.Float(readonly=True)
    mtr_p = fields.Float(readonly=True)
    mtr_s = fields.Float(readonly=True)
    mtr_cu = fields.Float(readonly=True)
    mtr_ni = fields.Float(readonly=True)
    mtr_cr = fields.Float(readonly=True)
    mtr_mo = fields.Float(readonly=True)
    mtr_n = fields.Float(readonly=True)
    mtr_yield_strength = fields.Float(readonly=True)
    mtr_tensile_strength = fields.Float(readonly=True)
    mtr_elongation = fields.Float(readonly=True)
    mtr_reduction_area = fields.Float(readonly=True)
    mtr_hardness = fields.Float(readonly=True)
    mtr_impact_test_temp = fields.Float(readonly=True)
    mtr_impact_coupon_size = fields.Char(readonly=True)
    mtr_impact_specimen_1 = fields.Float(readonly=True)
    mtr_impact_specimen_2 = fields.Float(readonly=True)
    mtr_impact_specimen_3 = fields.Float(readonly=True)
    mtr_impact_average = fields.Float(readonly=True)
    mtr_country_of_melt = fields.Char(readonly=True)
    mtr_country_of_manufacture = fields.Char(readonly=True)
    mtr_source_file = fields.Char(readonly=True)
    mtr_uploaded_at = fields.Datetime(readonly=True)

    def init(self):
        tools.drop_view_if_exists(self.env.cr, self._table)
        self.env.cr.execute(
            """
            CREATE OR REPLACE VIEW mtr_inventory_join_report AS (
                WITH latest_mtr AS (
                    SELECT
                        m.*,
                        row_number() OVER (
                            PARTITION BY m.heat_number
                            ORDER BY m.uploaded_at DESC NULLS LAST, m.id DESC
                        ) AS rn
                    FROM mtr_data m
                )
                SELECT
                    i.id AS id,
                    CASE WHEN m.id IS NULL THEN 'Missing MTR' ELSE 'Matched' END AS join_status,
                    i.entry_no AS inv_entry_no,
                    i.date AS inv_date,
                    i.location_code AS inv_location_code,
                    i.item_no AS inv_item_no,
                    i.quantity AS inv_quantity,
                    i.unit_of_measure_code AS inv_unit_of_measure_code,
                    i.document_no AS inv_document_no,
                    i.wsi_variant_code AS inv_wsi_variant_code,
                    i.dimensions AS inv_dimensions,
                    i.lot_number AS inv_lot_number,
                    i.heat_number AS inv_heat_number,
                    i.slab_number AS inv_slab_number,
                    i.internal_bin AS inv_internal_bin,
                    i.additional_notes AS inv_additional_notes,
                    i.cost_amount_actual AS inv_cost_amount_actual,
                    i.description_2 AS inv_description_2,
                    i.origin_code AS inv_origin_code,
                    i.picked AS inv_picked,
                    i.cutting_plan_no AS inv_cutting_plan_no,
                    i.image_path AS inv_image_path,
                    i.entry_type AS inv_entry_type,
                    i.document_type AS inv_document_type,
                    i.drawing AS inv_drawing,
                    i.yield_value AS inv_yield_value,
                    i.document_line_no AS inv_document_line_no,
                    i.revision AS inv_revision,
                    i.laser_quality AS inv_laser_quality,
                    i.unitcost_cwt AS inv_unitcost_cwt,
                    i.piece_no AS inv_piece_no,
                    i.variant_code AS inv_variant_code,
                    i.description AS inv_description,
                    i.return_reason_code AS inv_return_reason_code,
                    i.serial_no AS inv_serial_no,
                    i.package_no AS inv_package_no,
                    i.invoiced_quantity AS inv_invoiced_quantity,
                    i.inventory_by_location AS inv_inventory_by_location,
                    i.inventory AS inv_inventory,
                    i.expiration_date AS inv_expiration_date,
                    i.remaining_quantity AS inv_remaining_quantity,
                    i.shipped_qty_not_returned AS inv_shipped_qty_not_returned,
                    i.reserved_quantity AS inv_reserved_quantity,
                    i.qty_per_unit_of_measure AS inv_qty_per_unit_of_measure,
                    i.sales_amount_expected AS inv_sales_amount_expected,
                    i.sales_amount_actual AS inv_sales_amount_actual,
                    i.cost_amount_expected AS inv_cost_amount_expected,
                    i.cost_amount_non_invtbl AS inv_cost_amount_non_invtbl,
                    i.item_description AS inv_item_description,
                    i.cost_amount_expected_acy AS inv_cost_amount_expected_acy,
                    i.cost_amount_actual_acy AS inv_cost_amount_actual_acy,
                    i.completely_invoiced AS inv_completely_invoiced,
                    i.cost_amount_non_invtbl_acy AS inv_cost_amount_non_invtbl_acy,
                    i.assemble_to_order AS inv_assemble_to_order,
                    i.drop_shipment AS inv_drop_shipment,
                    i.open_flag AS inv_open_flag,
                    i.order_type AS inv_order_type,
                    i.order_no AS inv_order_no,
                    i.order_line_no AS inv_order_line_no,
                    i.prod_order_comp_line_no AS inv_prod_order_comp_line_no,
                    i.project_no AS inv_project_no,
                    i.project_task_no AS inv_project_task_no,
                    i.source_type AS inv_source_type,
                    i.source_no AS inv_source_no,
                    i.source_description AS inv_source_description,
                    i.source_order_no AS inv_source_order_no,
                    i.grade AS inv_grade,
                    i.weight AS inv_weight,
                    i.posting_date AS inv_posting_date,
                    i.country_of_melt AS inv_country_of_melt,
                    i.country_of_manufacture AS inv_country_of_manufacture,
                    i.source_file AS inv_source_file,
                    m.id AS mtr_id,
                    m.heat_number AS mtr_heat_number,
                    m.batch_number AS mtr_batch_number,
                    m.grade AS mtr_grade,
                    m.manufacturer AS mtr_manufacturer,
                    m.certificate_number AS mtr_certificate_number,
                    m.certificate_date AS mtr_certificate_date,
                    m.c_element AS mtr_c,
                    m.mn_element AS mtr_mn,
                    m.si_element AS mtr_si,
                    m.p_element AS mtr_p,
                    m.s_element AS mtr_s,
                    m.cu_element AS mtr_cu,
                    m.ni_element AS mtr_ni,
                    m.cr_element AS mtr_cr,
                    m.mo_element AS mtr_mo,
                    m.n_element AS mtr_n,
                    m.yield_strength AS mtr_yield_strength,
                    m.tensile_strength AS mtr_tensile_strength,
                    m.elongation AS mtr_elongation,
                    m.reduction_area AS mtr_reduction_area,
                    m.hardness AS mtr_hardness,
                    m.impact_test_temp AS mtr_impact_test_temp,
                    m.impact_coupon_size AS mtr_impact_coupon_size,
                    m.impact_specimen_1 AS mtr_impact_specimen_1,
                    m.impact_specimen_2 AS mtr_impact_specimen_2,
                    m.impact_specimen_3 AS mtr_impact_specimen_3,
                    m.impact_average AS mtr_impact_average,
                    m.country_of_melt AS mtr_country_of_melt,
                    m.country_of_manufacture AS mtr_country_of_manufacture,
                    m.source_file AS mtr_source_file,
                    m.uploaded_at AS mtr_uploaded_at
                FROM inventory i
                LEFT JOIN latest_mtr m
                    ON i.heat_number = m.heat_number
                   AND m.rn = 1
            )
            """
        )
